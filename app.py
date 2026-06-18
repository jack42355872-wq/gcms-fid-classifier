"""
GC-MS FID 成分分類網站
自動對 Shimadzu GC-MS 輸出的 xlsx 執行成分分類、顏色標記、統計加總
"""

import io
import os
import pandas as pd
from flask import Flask, request, send_file, render_template, jsonify
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

COLOR_MAP = {
    'n-alkane':     'FF8C00',
    'iso-alkane':   'FF6600',
    'cyclic':       'FFD700',
    'Alkene':       'ADD8E6',
    'Aromatic':     '92D050',
    'Polyaromatic': 'C5B0D5',
    'others':       'FA8072',
}

RT_C8_MINUS_MAX = 2.178
RT_C816_MAX     = 17.6



# ── PubChem Hybrid 查詢（僅對 others 觸發）────────────────────────────
_pubchem_cache: dict = {}   # in-memory cache，每次重啟清空

def _pubchem_lookup(name: str) -> str:
    """
    查 PubChem PUG REST API 取得分子式，依分子式二次分類。
    失敗時靜默回傳 'others'。
    """
    import urllib.request, urllib.parse, json as _json

    key = name.upper().strip()
    if key in _pubchem_cache:
        return _pubchem_cache[key]

    try:
        encoded = urllib.parse.quote(key)
        url = (f'https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/name/'
               f'{encoded}/property/MolecularFormula,IUPACName/JSON')
        with urllib.request.urlopen(url, timeout=5) as r:
            data = _json.loads(r.read())
        props = data['PropertyTable']['Properties'][0]
        formula = props.get('MolecularFormula', '')
        iupac   = props.get('IUPACName', '').upper()
        result  = _classify_by_formula(formula, iupac, key)
    except Exception:
        result = 'others'

    _pubchem_cache[key] = result
    return result


def _classify_by_formula(formula: str, iupac: str, original: str) -> str:
    """
    依分子式 CₙHₘ 推斷化合物類別。
    規則：
      - 含苯環特徵（H 數明顯不飽和且 IUPAC 含 benzene/phenyl/naphthalene）→ Aromatic/Polyaromatic
      - CₙH₂ₙ₋₆ (苯環) 或 IUPAC 含 benzene → Aromatic
      - IUPAC 含 naphthalene/pyrene/fluorene/anthracene → Polyaromatic
      - CₙH₂ₙ₊₂ 且含 cyclo → cyclic  (實際上 cycloalkane = CₙH₂ₙ，見下)
      - CₙH₂ₙ    → cyclic（若 iupac 含 cycl）或 Alkene
      - CₙH₂ₙ₊₂  → alkane；有 methyl branch → iso-alkane，否則 → n-alkane
    """
    import re
    m = re.match(r'^C(\d+)H(\d+)(?:O\d*)?(?:N\d*)?$', formula)
    if not m:
        return 'others'

    c = int(m.group(1))
    h = int(m.group(2))

    # 多環芳烴
    if any(x in iupac for x in ['NAPHTHALENE','PYRENE','FLUORENE',
                                  'ANTHRACENE','PHENANTHRENE','BIPHENYL']):
        return 'Polyaromatic'

    # 單環芳烴（苯環 CₙH₂ₙ₋₆）
    if h == 2 * c - 6 or any(x in iupac for x in ['BENZENE','TOLUENE',
                                                     'XYLENE','PHENYL',
                                                     'STYRENE','CUMENE']):
        return 'Aromatic'

    # Cycloalkane CₙH₂ₙ + cycl in name
    if h == 2 * c and 'CYCL' in iupac:
        return 'cyclic'

    # Alkene CₙH₂ₙ（不含 cyclo）
    if h == 2 * c:
        return 'Alkene'

    # Alkane CₙH₂ₙ₊₂
    if h == 2 * c + 2:
        if 'METHYL' in iupac or 'ETHYL' in iupac:
            # 直鏈 ethyl ≠ iso，再確認是否 branched
            if re.search(r'\d+-methyl|\d+-ethyl|\d+-propyl', iupac):
                return 'iso-alkane'
        return 'n-alkane'

    return 'others'

def auto_classify(name: str) -> str:
    n = name.upper().strip()
    if any(x in n for x in ['NAPHTHALENE','ANTHRACENE','PHENANTHRENE',
                              'PYRENE','FLUORENE','BIPHENYL','ACENAPHTHY',
                              'INDENE','TETRAHYDRONAPHTHALENE','TETRALIN']):
        return 'Polyaromatic'
    if any(x in n for x in ['BENZENE','TOLUENE','XYLENE','STYRENE',
                              'INDAN','PHENYL','CUMENE']):
        return 'Aromatic'
    if any(x in n for x in ['CYCLOPENTANE','CYCLOHEXANE','CYCLOPROPANE',
                              'CYCLOBUTANE','CYCLOHEPTANE','CYCLOOCTANE']):
        return 'cyclic'
    if '-ENE' in n or any(x in n for x in ['PROPENE','BUTENE','PENTENE',
                                             'HEXENE','HEPTENE','OCTENE',
                                             'NONENE','DECENE']):
        return 'Alkene'
    straight = ['HEXANE','HEPTANE','OCTANE','NONANE','DECANE',
                'UNDECANE','DODECANE','TRIDECANE','TETRADECANE',
                'PENTADECANE','HEXADECANE','HEPTADECANE','OCTADECANE']
    if any(x in n for x in straight):
        if '2-METHYL' in n or '3-METHYL' in n:
            return 'iso-alkane'
        return 'n-alkane'
    if any(x in n for x in ['ISOPENTANE','ISOBUTANE','2-METHYL','3-METHYL']):
        if 'ANE' in n:
            return 'iso-alkane'
    return 'others'


def _pubchem_fallback(name: str) -> str:
    """先 auto_classify，若結果是 others 才查 PubChem。"""
    result = auto_classify(name)
    if result == 'others':
        result = _pubchem_lookup(name)
    return result


def process_gcms_file(file_bytes: bytes, original_filename: str):
    # Step 1: 動態定位 header row
    df_raw = pd.read_excel(io.BytesIO(file_bytes), sheet_name='Sheet1', header=None)
    header_row_idx = None
    for i, row in df_raw.iterrows():
        vals = [str(v) for v in row.values]
        if 'Peak#' in vals or 'Ret.Time' in vals:
            header_row_idx = i
            break
    if header_row_idx is None:
        raise ValueError("找不到 Peak Table header")

    data = df_raw.iloc[header_row_idx:].copy()
    data.columns = data.iloc[0]
    data = data.iloc[1:].reset_index(drop=True)
    data.columns = [str(c).strip().replace('\xa0', '') for c in data.columns]

    # Step 2: 清理
    data['Ret.Time'] = pd.to_numeric(data['Ret.Time'], errors='coerce')
    data['Conc.']    = pd.to_numeric(data['Conc.'],    errors='coerce')
    data['Name']     = data['Name'].astype(str).str.strip() if 'Name' in data.columns else ''
    if 'Category' in data.columns:
        data['Category'] = data['Category'].astype(str).str.strip().str.replace('\xa0','',regex=False)
    else:
        data['Category'] = ''

    # Step 3: 自動分類
    mask_empty = data['Category'].isin(['nan','','NaN']) | data['Category'].isna()
    data.loc[mask_empty, 'Category'] = data.loc[mask_empty, 'Name'].apply(_pubchem_fallback)
    auto_classified_count = int(mask_empty.sum())

    # Step 4: 計算統計（在操作 Excel 前完成）
    mask_c8m  = data['Ret.Time'] < RT_C8_MINUS_MAX
    mask_c816 = (data['Ret.Time'] >= RT_C8_MINUS_MAX) & (data['Ret.Time'] <= RT_C816_MAX)
    mask_c16p = data['Ret.Time'] > RT_C816_MAX

    c8minus    = round(data.loc[mask_c8m,  'Conc.'].sum(), 2)
    c16plus    = round(data.loc[mask_c16p, 'Conc.'].sum(), 2)
    filt816    = data[mask_c816]
    c816_total = round(filt816['Conc.'].sum(), 2)
    sums_816   = filt816.groupby('Category', dropna=True)['Conc.'].sum()

    n_val     = sums_816.get('n-alkane',   0.0)
    iso_val   = sums_816.get('iso-alkane', 0.0)
    i_n_ratio = round(iso_val / n_val, 2) if n_val > 0 else 0.0

    grand_total  = c8minus + c816_total + c16plus
    all_conc_sum = round(data['Conc.'].sum(), 2)
    error        = abs(grand_total - all_conc_sum)
    passed       = error < 0.01
    others_816   = filt816[filt816['Category']=='others'][['Ret.Time','Name','Conc.']].to_dict('records')

    # Step 5: 開啟 Excel
    wb = load_workbook(io.BytesIO(file_bytes))
    ws = wb['Sheet1']

    # 確認是否已有 Summary 表格
    summary_header_row = None
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == 'C8-':
                summary_header_row = cell.row
                break
        if summary_header_row:
            break

    rows_inserted = 0
    if summary_header_row is None:
        insert_at = header_row_idx + 1  # 0-based → 1-based
        ws.insert_rows(insert_at, amount=4)
        rows_inserted = 4
        summary_header_row = insert_at

    # DATA_START_ROW 同步更新
    DATA_START_ROW = header_row_idx + 2 + rows_inserted

    # Step 6: 找/建 Category、Distribution 欄
    peak_header_excel_row = header_row_idx + 1 + rows_inserted
    cat_col_excel = None
    for cell in ws[peak_header_excel_row]:
        v = str(cell.value).strip().replace('\xa0','') if cell.value else ''
        if v == 'Category':
            cat_col_excel = cell.column

    if cat_col_excel is None:
        cat_col_excel = ws.max_column + 1
        ws.cell(row=peak_header_excel_row, column=cat_col_excel,
                value='Category').font = Font(bold=True)

    dist_col_excel = cat_col_excel + 1
    if ws.cell(row=peak_header_excel_row, column=dist_col_excel).value != 'Distribution':
        ws.cell(row=peak_header_excel_row, column=dist_col_excel,
                value='Distribution').font = Font(bold=True)

    # Step 7: 上色（只染 Peak Table 資料列）
    for idx, row_data in data.iterrows():
        excel_row = DATA_START_ROW + idx
        rt       = row_data['Ret.Time']
        category = row_data['Category']

        ws.cell(row=excel_row, column=cat_col_excel).value = category
        if pd.notna(rt):
            if rt < RT_C8_MINUS_MAX:
                dist = 'C8-'
            elif rt <= RT_C816_MAX:
                dist = 'C8-16'
            else:
                dist = 'C16+'
            ws.cell(row=excel_row, column=dist_col_excel).value = dist

        if pd.notna(rt) and RT_C8_MINUS_MAX <= rt <= RT_C816_MAX and category in COLOR_MAP:
            fill = PatternFill(start_color=COLOR_MAP[category],
                               end_color=COLOR_MAP[category], fill_type='solid')
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=excel_row, column=col_idx).fill = fill

    # Step 8: 寫 Summary 表格
    SUMMARY_VALUES_ROW = summary_header_row + 1
    summary_labels = ['C8-','C8-16','C16+','i/n',
                      'alkene','aromatic','cyclic','iso','n-alkane','others','poly']
    header_colors  = ['FFD966','FFD966','FFD966','FFD966',
                      'ADD8E6','92D050','FFD700','FF6600','FF8C00','FA8072','C5B0D5']

    col_map = {}
    for cell in ws[summary_header_row]:
        if cell.value:
            col_map[str(cell.value).strip()] = cell.column

    if not col_map:
        for col_offset, (label, color) in enumerate(zip(summary_labels, header_colors)):
            cell = ws.cell(row=summary_header_row, column=col_offset+1)
            cell.value     = label
            cell.font      = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill      = PatternFill(start_color=color, end_color=color, fill_type='solid')
            col_map[label] = col_offset + 1

    value_map = {
        'C8-':      c8minus,
        'C8-16':    c816_total,
        'C16+':     c16plus,
        'i/n':      i_n_ratio,
        'alkene':   round(sums_816.get('Alkene',       0.0), 2),
        'aromatic': round(sums_816.get('Aromatic',     0.0), 2),
        'cyclic':   round(sums_816.get('cyclic',       0.0), 2),
        'iso':      round(sums_816.get('iso-alkane',   0.0), 2),
        'n-alkane': round(sums_816.get('n-alkane',     0.0), 2),
        'others':   round(sums_816.get('others',       0.0), 2),
        'poly':     round(sums_816.get('Polyaromatic', 0.0), 2),
    }
    for label, val in value_map.items():
        if label in col_map:
            cell = ws.cell(row=SUMMARY_VALUES_ROW, column=col_map[label])
            cell.value         = val
            cell.number_format = '0.00'
            cell.alignment     = Alignment(horizontal='center')
            cell.font          = Font(bold=True)
            cell.fill          = PatternFill(start_color='FFF2CC',
                                             end_color='FFF2CC', fill_type='solid')

    # Step 9: BarChart（放在 Summary 右側 N 欄）
    chart_data_col = len(summary_labels) + 3
    chart_labels = ['Aromatic','Polyaromatic','n-alkane','iso-alkane','cyclic','Alkene','others']
    chart_values = [
        round(sums_816.get('Aromatic',     0.0), 2),
        round(sums_816.get('Polyaromatic', 0.0), 2),
        round(sums_816.get('n-alkane',     0.0), 2),
        round(sums_816.get('iso-alkane',   0.0), 2),
        round(sums_816.get('cyclic',       0.0), 2),
        round(sums_816.get('Alkene',       0.0), 2),
        round(sums_816.get('others',       0.0), 2),
    ]

    ws.cell(row=summary_header_row,   column=chart_data_col).value = 'Category'
    ws.cell(row=summary_header_row,   column=chart_data_col+1).value = 'Conc%'
    for i, (lbl, val) in enumerate(zip(chart_labels, chart_values)):
        ws.cell(row=summary_header_row+1+i, column=chart_data_col).value   = lbl
        ws.cell(row=summary_header_row+1+i, column=chart_data_col+1).value = val

    from openpyxl.chart.label import DataLabel, DataLabelList
    from openpyxl.chart.layout import Layout

    chart = BarChart()
    chart.type       = 'col'
    chart.grouping   = 'clustered'
    chart.title      = 'C8-16 Category Distribution (%)'
    chart.y_axis.title  = 'Conc. %'
    chart.x_axis.title  = 'Category'
    chart.width      = 22
    chart.height     = 14
    chart.y_axis.numFmt = '0.00'
    chart.y_axis.majorGridlines = None

    data_ref = Reference(ws,
                         min_col=chart_data_col+1,
                         min_row=summary_header_row,
                         max_row=summary_header_row+len(chart_labels))
    cats_ref = Reference(ws,
                         min_col=chart_data_col,
                         min_row=summary_header_row+1,
                         max_row=summary_header_row+len(chart_labels))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    bar_colors = ['92D050','C5B0D5','FF8C00','FF6600','FFD700','ADD8E6','FA8072']
    for idx, color in enumerate(bar_colors):
        pt = DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = color
        chart.series[0].dPt.append(pt)

    # 加 data labels（顯示數值在 bar 上方）
    chart.series[0].dLbls = DataLabelList()
    chart.series[0].dLbls.showVal   = True
    chart.series[0].dLbls.showLegendKey = False
    chart.series[0].dLbls.showCatName   = False
    chart.series[0].dLbls.numFmt        = '0.00'

    ws.add_chart(chart, f'N{summary_header_row}')

    # Step 10: Validation Report（Peak Table 最後一列之後）
    report_start_row = DATA_START_ROW + len(data) + 2

    title_cell = ws.cell(row=report_start_row, column=1)
    title_cell.value = 'Validation Report'
    title_cell.fill  = PatternFill(start_color='2B6CB0', end_color='2B6CB0', fill_type='solid')
    title_cell.font  = Font(bold=True, color='FFFFFF', size=11)

    report_rows = [
        ('C8-  (conc%)',         c8minus),
        ('C8-16  (conc%)',       c816_total),
        ('C16+  (conc%)',        c16plus),
        ('  Aromatic',           round(sums_816.get('Aromatic',     0.0), 2)),
        ('  Polyaromatic',       round(sums_816.get('Polyaromatic', 0.0), 2)),
        ('  n-alkane',           round(sums_816.get('n-alkane',     0.0), 2)),
        ('  iso-alkane',         round(sums_816.get('iso-alkane',   0.0), 2)),
        ('  cyclic',             round(sums_816.get('cyclic',       0.0), 2)),
        ('  Alkene',             round(sums_816.get('Alkene',       0.0), 2)),
        ('  others',             round(sums_816.get('others',       0.0), 2)),
        ('Total check (sum)',    grand_total),
        ('All Conc. sum',        all_conc_sum),
        ('delta error',          round(error, 4)),
        ('PASS / FAIL',          'PASS' if passed else f'FAIL (delta={round(error,4)})'),
        ('Auto-classified rows', auto_classified_count),
    ]

    for offset, (label, value) in enumerate(report_rows):
        r = report_start_row + 1 + offset
        ws.cell(row=r, column=1).value = label
        ws.cell(row=r, column=2).value = value
        ws.cell(row=r, column=1).font  = Font(size=10)
        ws.cell(row=r, column=2).font  = Font(bold=True, size=10)
        if label == 'PASS / FAIL':
            bg = 'C6EFCE' if passed else 'FFC7CE'
            ws.cell(row=r, column=2).fill = PatternFill(
                start_color=bg, end_color=bg, fill_type='solid')

    if others_816:
        others_start = report_start_row + 1 + len(report_rows) + 1
        ws.cell(row=others_start, column=1).value = 'Unclassified (others) in C8-16'
        ws.cell(row=others_start, column=1).font  = Font(bold=True, color='FF0000')
        for i, rec in enumerate(others_816):
            r = others_start + 1 + i
            ws.cell(row=r, column=1).value = rec.get('Name', '')
            ws.cell(row=r, column=2).value = rec.get('Ret.Time', '')
            ws.cell(row=r, column=3).value = rec.get('Conc.', '')

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    validation_report = {
        'c8minus':         c8minus,
        'c816_total':      c816_total,
        'c16plus':         c16plus,
        'grand_total':     grand_total,
        'all_conc_sum':    all_conc_sum,
        'error':           round(error, 4),
        'passed':          passed,
        'auto_classified': auto_classified_count,
        'others_816':      others_816,
        'sums_816':        {k: round(v, 2) for k, v in sums_816.items()},
    }
    return out.read(), validation_report


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/classify', methods=['POST'])
def classify():
    if 'file' not in request.files:
        return jsonify({'error': 'no file'}), 400
    f = request.files['file']
    if not f.filename.endswith('.xlsx'):
        return jsonify({'error': 'xlsx only'}), 400

    file_bytes = f.read()
    original_filename = secure_filename(f.filename)

    try:
        result_bytes, report = process_gcms_file(file_bytes, original_filename)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

    stem = original_filename.rsplit('.', 1)[0]
    out_filename = f"{stem}_classified.xlsx"

    response = send_file(
        io.BytesIO(result_bytes),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=out_filename,
    )
    response.headers['X-C8minus']      = str(report['c8minus'])
    response.headers['X-C816total']    = str(report['c816_total'])
    response.headers['X-C16plus']      = str(report['c16plus'])
    response.headers['X-GrandTotal']   = str(report['grand_total'])
    response.headers['X-Passed']       = str(report['passed'])
    response.headers['X-AutoClass']    = str(report['auto_classified'])
    s816 = report['sums_816']
    response.headers['X-Aromatic']     = str(round(s816.get('Aromatic',     0.0), 2))
    response.headers['X-Polyaromatic'] = str(round(s816.get('Polyaromatic', 0.0), 2))
    response.headers['X-Nalkane']      = str(round(s816.get('n-alkane',     0.0), 2))
    response.headers['X-Isoalkane']    = str(round(s816.get('iso-alkane',   0.0), 2))
    response.headers['X-Cyclic']       = str(round(s816.get('cyclic',       0.0), 2))
    response.headers['X-Alkene']       = str(round(s816.get('Alkene',       0.0), 2))
    response.headers['X-Others']       = str(round(s816.get('others',       0.0), 2))
    n_val   = s816.get('n-alkane',   0.0)
    iso_val = s816.get('iso-alkane', 0.0)
    response.headers['X-InRatio'] = str(round(iso_val / n_val, 2) if n_val > 0 else 0.0)
    return response


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)

